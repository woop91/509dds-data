"""
extract-source-spreadsheets.py

Reads the 5 xlsx files imported from the desktop archive (data/ssa/{staffing,
claims,allowance-rates}/*.source.xlsx) and emits paired JSON extracts in
long-form {scope, code, year, metric, value} shape for downstream Supabase
insertion.

Run from repo root:
    python scripts/extract-source-spreadsheets.py

Idempotent — overwrites the .json files next to each .source.xlsx.
"""
import json
import os
import re
from datetime import datetime, timezone
from typing import Iterable

import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def write_json(path: str, payload: dict) -> None:
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.write("\n")
    print(f"  wrote {os.path.relpath(path, REPO)}  ({len(payload.get('rows', []))} rows)")


def coerce_number(v):
    """Return float for numeric, None for redacted markers ('2/', etc), None for blanks."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "2/", "3/", "*", "N/A", "n/a"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


# SSA regional codes — first column entries that name a region (vs an individual state).
SSA_REGIONS = {
    "BOSTON", "BOS",
    "NEW YORK", "NY-REGION",
    "PHILADELPHIA", "PHI",
    "ATLANTA", "ATL",
    "CHICAGO", "CHI",
    "DALLAS", "DAL",
    "KANSAS CITY", "KC",
    "DENVER", "DEN",
    "SAN FRANCISCO", "SF",
    "SEATTLE", "SEA",
    "NATIONAL", "NATIONAL - DDS ONLY", "GRAND TOTAL",
}


def classify_scope(code: str) -> str:
    c = code.upper().strip()
    if c.startswith("NATIONAL") or c == "GRAND TOTAL":
        return "national"
    if c in SSA_REGIONS:
        return "region"
    return "state"


def extract_examiners(path: str) -> dict:
    """examiners by state FY2002-2018 — wide format, one row per state, one column per FY."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # row 0 = title, row 1 = year headers, rows 2+ = data
    years = [int(y) for y in rows[1][1:] if isinstance(y, int)]
    out_rows = []
    for r in rows[2:]:
        code = r[0]
        if not code:
            continue
        code = str(code).strip()
        if not code:
            continue
        scope = classify_scope(code)
        for i, y in enumerate(years, start=1):
            v = coerce_number(r[i])
            if v is None:
                continue
            out_rows.append({
                "scope": scope,
                "code": code,
                "year": y,
                "metric": "dds_examiners_count",
                "value": v,
            })
    return {
        "schema": "long",
        "metric_family": "dds_examiners_by_state",
        "source_file": "data/ssa/staffing/dds-examiners-by-state-fy2002-2018.source.xlsx",
        "extracted_at": now_utc(),
        "fy_range": [min(years), max(years)],
        "rows": out_rows,
    }


def extract_employees(path: str) -> dict:
    """all DDS employees by state FY2010-2023 — hierarchical: col 1 = region/Grand Total
    (with col 2 = 'Total'); col 2 = individual state when col 1 is empty."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[2]
    years = []
    year_cols = []
    for ci, cell in enumerate(header_row):
        if isinstance(cell, str):
            m = re.match(r"^FY\s+(\d{4})$", cell.strip())
            if m:
                years.append(int(m.group(1)))
                year_cols.append(ci)
    out_rows = []
    for r in rows[4:]:
        col1 = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        col2 = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        if col1:
            # Region or Grand Total row (col 2 is typically 'Total')
            code = col1
            scope = classify_scope(code)
        elif col2 and col2.upper() != "TOTAL":
            # State row nested under previous region
            code = col2
            scope = "state"
        else:
            continue
        for y, ci in zip(years, year_cols):
            v = coerce_number(r[ci])
            if v is None:
                continue
            out_rows.append({
                "scope": scope,
                "code": code,
                "year": y,
                "metric": "total_dds_employees",
                "value": v,
            })
    return {
        "schema": "long",
        "metric_family": "total_dds_employees_by_state",
        "source_file": "data/ssa/staffing/dds-employees-by-state-fy2010-2023.source.xlsx",
        "extracted_at": now_utc(),
        "fy_range": [min(years), max(years)] if years else None,
        "rows": out_rows,
    }


def extract_claims_by_title(path: str) -> dict:
    """initial claims by Title FY2001-2019 — long-ish, FY rows × {SSDI, SSI, Concurrent, Total}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # row 1 is column headers
    headers = [str(c or "").strip() for c in rows[1]]
    # Map columns to canonical metric names
    title_map = {
        "Social Security Disabililty Insurance": "initial_claims_ssdi",  # sic in source
        "Social Security Disability Insurance": "initial_claims_ssdi",
        "Supplement Secuirty Income": "initial_claims_ssi",  # sic in source
        "Supplemental Security Income": "initial_claims_ssi",
        "Concurrent (filing for both SSDI and SSI)": "initial_claims_concurrent",
        "Total": "initial_claims_total",
    }
    metric_cols = {}
    for ci, h in enumerate(headers):
        if h in title_map:
            metric_cols[ci] = title_map[h]
    out_rows = []
    for r in rows[2:]:
        if not r or not r[0]:
            continue
        # FY may be int or str
        fy_raw = r[0]
        try:
            fy = int(fy_raw)
        except (TypeError, ValueError):
            continue
        if fy < 1990 or fy > 2050:
            continue
        for ci, metric in metric_cols.items():
            v = coerce_number(r[ci])
            if v is None:
                continue
            out_rows.append({
                "scope": "national",
                "code": "US",
                "year": fy,
                "metric": metric,
                "value": v,
            })
    return {
        "schema": "long",
        "metric_family": "initial_claims_by_title",
        "source_file": "data/ssa/claims/initial-claims-by-title-fy2001-2019.source.xlsx",
        "extracted_at": now_utc(),
        "rows": out_rows,
    }


def extract_allowance_rate_sheet(ws) -> tuple[int | None, list[dict]]:
    """Pull (year, [rows]) from one allowance-rate sheet. Year detected from preamble."""
    rows = list(ws.iter_rows(values_only=True))
    # Find 'Fiscal Year YYYY' in the first 10 rows
    year = None
    for r in rows[:10]:
        for cell in r:
            if isinstance(cell, str):
                m = re.search(r"Fiscal Year\s+(\d{4})", cell)
                if m:
                    year = int(m.group(1))
                    break
        if year:
            break
    # Find the header row containing 'INITIAL' and 'RECONSIDERATION'
    header_row_idx = None
    for ri, r in enumerate(rows[:15]):
        cells = [str(c).upper() if c else "" for c in r]
        if "INITIAL" in cells and "RECONSIDERATION" in cells:
            header_row_idx = ri
            break
    if header_row_idx is None or year is None:
        return year, []
    initial_col = next(i for i, c in enumerate(rows[header_row_idx]) if str(c or "").upper() == "INITIAL")
    recon_col = next(i for i, c in enumerate(rows[header_row_idx]) if str(c or "").upper() == "RECONSIDERATION")
    out = []
    for r in rows[header_row_idx + 1:]:
        if not r or not r[0]:
            continue
        code = str(r[0]).strip()
        if not code:
            continue
        # Strip footnote markers like 'PR3/', 'PR2/', 'NH3/'
        code = re.sub(r"\d/$", "", code).strip()
        scope = classify_scope(code)
        initial = coerce_number(r[initial_col])
        recon = coerce_number(r[recon_col])
        if initial is not None:
            out.append({
                "scope": scope, "code": code, "year": year,
                "metric": "allowance_rate_initial_pct", "value": initial,
            })
        if recon is not None:
            out.append({
                "scope": scope, "code": code, "year": year,
                "metric": "allowance_rate_recon_pct", "value": recon,
            })
    return year, out


def extract_allowance_rates_2018_2024(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    all_rows = []
    years_seen = []
    for sname in wb.sheetnames:
        year, rows = extract_allowance_rate_sheet(wb[sname])
        if year and rows:
            all_rows.extend(rows)
            years_seen.append(year)
    return {
        "schema": "long",
        "metric_family": "allowance_rates_by_state",
        "source_file": "data/ssa/allowance-rates/2018-2024.source.xlsx",
        "extracted_at": now_utc(),
        "fy_range": [min(years_seen), max(years_seen)] if years_seen else None,
        "rows": all_rows,
    }


def extract_allowance_rate_fy2019_foia(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    # Single sheet (named 'FY2018' but contains FY2019 data per preamble)
    year, rows = extract_allowance_rate_sheet(wb[wb.sheetnames[0]])
    return {
        "schema": "long",
        "metric_family": "allowance_rates_by_state_foia",
        "source_file": "data/ssa/allowance-rates/fy2019-foia-reading-room.source.xlsx",
        "provenance": "SSA FOIA Reading Room",
        "extracted_at": now_utc(),
        "fy_range": [year, year] if year else None,
        "rows": rows,
    }


def main():
    base = os.path.join(REPO, "data", "ssa")
    targets = [
        (
            os.path.join(base, "staffing", "dds-examiners-by-state-fy2002-2018.source.xlsx"),
            os.path.join(base, "staffing", "dds-examiners-by-state-fy2002-2018.json"),
            extract_examiners,
        ),
        (
            os.path.join(base, "staffing", "dds-employees-by-state-fy2010-2023.source.xlsx"),
            os.path.join(base, "staffing", "dds-employees-by-state-fy2010-2023.json"),
            extract_employees,
        ),
        (
            os.path.join(base, "claims", "initial-claims-by-title-fy2001-2019.source.xlsx"),
            os.path.join(base, "claims", "initial-claims-by-title-fy2001-2019.json"),
            extract_claims_by_title,
        ),
        (
            os.path.join(base, "allowance-rates", "2018-2024.source.xlsx"),
            os.path.join(base, "allowance-rates", "allowance-rates-by-state-2018-2024.json"),
            extract_allowance_rates_2018_2024,
        ),
        (
            os.path.join(base, "allowance-rates", "fy2019-foia-reading-room.source.xlsx"),
            os.path.join(base, "allowance-rates", "fy2019-foia-reading-room.json"),
            extract_allowance_rate_fy2019_foia,
        ),
    ]
    for src, dst, fn in targets:
        print(f"=> {os.path.relpath(src, REPO)}")
        payload = fn(src)
        write_json(dst, payload)


if __name__ == "__main__":
    main()
